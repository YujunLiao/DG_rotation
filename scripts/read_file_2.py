import os
import numpy
import re
from openpyxl import Workbook
import math

wb = Workbook()
# grab the active worksheet
ws = wb.active

path = '/home/lyj/Files/project/pycharm/DG_rotation/trainer_utils/output_manager/output_file/'\
+'era/DG_rotation/caffenet/'
path_of_different_parameter_pairs = os.listdir(path)
path_of_different_parameter_pairs.sort()


row = 1
for parameter_pair_path in path_of_different_parameter_pairs:
    if parameter_pair_path != '0.4_0.4':
        continue
    column = 0
    ws[chr(ord('A') + 1) + str(row)] = parameter_pair_path
    row += 1
    ws[chr(ord('A')) + str(row)] = 'epoch'
    ws[chr(ord('A') + 1) + str(row)] = 'self-supervised'
    ws[chr(ord('A') + 2) + str(row)] = 'classification'
    row += 1


    origin_file= 'original_record_2'
    if not os.path.exists(path + parameter_pair_path + '/' + origin_file):
        print('error')
    # origin_file_words = origin_file.split('_')
    # abbreviation = origin_file_words[0][0] + origin_file_words[1][0]
    # ws[chr(ord('A') + column) + str(row)] = abbreviation
    # row += 1
    # print(abbreviation, origin_file_words)
    repeat_number = 0
    lines = open(path + parameter_pair_path + '/' + origin_file, 'r')
    for line in lines:
        if 'Start' in line:
            repeat_number += 1

            if repeat_number == 4:
                row = 3
                column += 4
                repeat_number = 1

        if repeat_number == 1:
            words = line.split('=')
            # if words[0] == 'target':
            #     print(words)
            if len(words) >1:
                index = words.index("'', target")
                target_domain = words[index+1].split("'")[1]
                ws[chr(ord('A') + column+1) + str(row)] = target_domain
                row += 1

        words = line.split(' ')
        if words[0]=='current':
            epoch = int(words[2][:-1])
            ws[chr(ord('A') + column) + str(row)] = int(words[2][:-1])


        if words[0]=='Accuracies' and words[2]=='test:':
            ws[chr(ord('A') + column+1) + str(row)] = float(words[5][:-1])
            ws[chr(ord('A') + column + 2) + str(row)] = float(words[8][:-1])
            # print(words, words[8])
            row += 1

            if epoch == 29:
                # pass
                row+=1



    row += 3
    print(parameter_pair_path)
print(path)
wb.save("./DG_rotation_caffenet_PACS_0.4_0.4.xlsx")