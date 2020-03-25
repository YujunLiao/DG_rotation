import os
import numpy
import re
from openpyxl import Workbook
import math

wb = Workbook()
# grab the active worksheet
ws = wb.active

path = '/home/lyj/Files/project/pycharm/DG_rotation/trainer_utils/output_manager/output_file/'\
+'era/DA_rotation/resnet50/'
path_of_different_parameter_pairs = os.listdir(path)
path_of_different_parameter_pairs.sort()


row = 1
for parameter_pair_path in path_of_different_parameter_pairs:
    path_of_different_domain_pairs= os.listdir(path + parameter_pair_path)
    path_of_different_domain_pairs.sort()
    # files = ['cartoon_art_painting','art_painting_cartoon',  'art_painting_sketch', 'art_painting_photo', ]
    column = 0
    ws[chr(ord('A') + 1) + str(row)] = parameter_pair_path
    row += 1
    for domain_pair in path_of_different_domain_pairs: #遍历文件夹
        if domain_pair=='original_record':
            continue
        domain_pair_words = domain_pair.split('_')
        abbreviation = domain_pair_words[0][0] + domain_pair_words[1][0]
        ws[chr(ord('A') + column) + str(row)] = abbreviation
        row += 1
        print(abbreviation, domain_pair_words)
        lines = open(path + parameter_pair_path + '/' + domain_pair, 'r')
        for line in lines:
            words = line.split(' ')
            if words[0]=='Accuracy':
                ws[chr(ord('A') + column) + str(row)] = words[11].split(':')[1][:-1]
                row += 1
        row -= 4
        column += 1

    row += 6

    print(parameter_pair_path)
print(path)
wb.save("./result.xlsx")